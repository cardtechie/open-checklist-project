import argparse
import re
import uuid
from pathlib import Path

import requests
import yaml
from openpyxl import load_workbook


DATA_DIR = Path("data") / "baseball"


PARALLEL_PLANS = {
    "2025-26-topps-chrome-platinum": {
        "source_url": "https://baseballcardpedia.com/index.php/2025-26_Topps_Chrome_Platinum#Checklist",
        "sections": {
            "Base": [
                ("Refractor", None),
                ("Prism Refractor", None),
                ("Blue Prism Refractor", None),
                ("Red Prism Refractor", None),
                ("Gold Prism Refractor", None),
                ("Topps Refractor", 499),
                ("Vibration Refractor", 250),
                ("Blue Mini-Diamond Refractor", 199),
                ("Speckle Refractor", 150),
                ("Blue Vibration Refractor", 150),
                ("Blue Lava Refractor", 100),
                ("Platinum Toile Cream/Fuchsia Shimmer Refractor", 100),
                ("Green Wave Refractor", 99),
                ("Platinum Toile White/Green Refractor", 99),
                ("Green Vibration Refractor", 99),
                ("Platinum Toile Cream/Rose Gold Refractor", 75),
                ("Rose Gold Refractor", 75),
                ("Rose Gold Mini-Diamond Refractor", 75),
                ("Diamond Etch Refractor", 55),
                ("Gold Refractor", 50),
                ("Gold Wave Refractor", 50),
                ("Platinum Toile Cream/Gold Refractor", 50),
                ("Gold Vibration Refractor", 50),
                ("Platinum Toile White/Orange Refractor", 25),
                ("Orange Refractor", 25),
                ("Orange Wave Refractor", 25),
                ("Orange Vibration Refractor", 25),
                ("Black Refractor", 10),
                ("Platinum Toile Cream/Black Refractor", 10),
                ("Black Vibration Refractor", 10),
                ("Red Refractor", 5),
                ("Platinum Toile Cream/Red Refractor", 5),
                ("Red Lava Refractor", 5),
                ("Red Vibration Refractor", 5),
                ("SuperFractor", 1),
            ],
            "1955 World Series": [
                ("Green Refractor", 99),
                ("Gold Refractor", 50),
                ("Orange Refractor", 25),
                ("Black Refractor", 10),
                ("Red Refractor", 5),
                ("SuperFractor", 1),
            ],
            "1955 Topps Rails And Sails": [
                ("Green Refractor", 99),
                ("Gold Refractor", 50),
                ("Orange Refractor", 25),
                ("Black Refractor", 10),
                ("Red Refractor", 5),
                ("SuperFractor", 1),
            ],
            "1955 Topps Doubleheaders": [
                ("Green Refractor", 99),
                ("Gold Refractor", 50),
                ("Orange Refractor", 25),
                ("Black Refractor", 10),
                ("Red Refractor", 5),
                ("SuperFractor", 1),
            ],
            "1955 Cards That Never Were": [("SuperFractor", 1)],
            "Base - City Variations": [("SuperFractor", 1)],
            "Chrome Platinum Autographs": [
                ("Refractor", 199),
                ("Aqua Refractor", 150),
                ("Blue Prism Refractor", 99),
                ("Platinum Toile Cream/Blue Refractor", 99),
                ("Green Mini-Diamond Refractor", 75),
                ("Gold Refractor", 50),
                ("Orange Refractor", 25),
                ("Pink Refractor", 15),
                ("Black Refractor", 10),
                ("Platinum Toile Cream/Red Refractor", 5),
                ("Red Refractor", 5),
                ("SuperFractor", 1),
            ],
            "1955 Topps City Variations Autographs": [("SuperFractor", 1)],
        },
    },
    "2025-26-topps-definitive-collection": {
        "source_url": "https://baseballcardpedia.com/index.php/2025-26_Topps_Definitive_Collection#Checklist",
        "sections": {
            "Jumbo Relic Collection": [("Orange", 25), ("Pink", 15), ("Black", 10), ("Red", 5), ("Gold", 1)],
            "Definitive Helmet Collection": [("Red", 5), ("MLB Logo Gold", 1), ("Rawlings Authentic Tag Gold", 1), ("Rawlings Logo Gold", 1), ("Player Number Gold", 1)],
            "Protector At The Plate": [("Black", 10), ("Red", 5), ("Gold", 1)],
            "Quad Patch Collection": [("Gold", 1)],
            "Definitive Rookie Autographs": [("Orange", 25), ("Black", 10), ("Red", 5), ("Gold", 1)],
            "Framed Autograph Collection": [("Black", 10), ("Red", 5), ("Gold", 1)],
            "Legendary Autograph Collection": [("Black", 10), ("Red", 5), ("Gold", 1)],
            "Defining Images Autographs": [("Black", 10), ("Red", 5), ("Gold", 1)],
            "Dual Autographs": [("Black", 10), ("Red", 5), ("Gold", 1)],
            "Definitive Triple Autographs": [("Black", 10), ("Red", 5), ("Gold", 1)],
            "Definitive Quad Autographs": [("Gold", 1)],
            "Dual World Series Autograph Collection": [("Black", 10), ("Red", 5), ("Gold", 1)],
            "Base Autographed Relics": [("Orange", 25), ("Black", 10), ("Red", 5), ("Jersey Button Gold", 1), ("Laundry Tag Gold", 1), ("Brand Logo Gold", 1), ("MLB Logo Gold", 1)],
            "Definitive Autographed Relics": [("Green", 10), ("Purple", 5), ("Red", 1)],
            "Definitive Rookie Patch Autographs": [("Orange", 25), ("Black", 10), ("Red", 5), ("Gold", 1)],
            "Definitive Autographed First Batting Gloves": [("Red", 5), ("Gold", 1)],
            "Autographed Patch Booklets": [("Red", 1)],
            "Framed Autograph Patches": [("Black", 10), ("Red", 5), ("Gold", 1)],
            "Definitive Helmet Collection Autographs": [("Gold", 1)],
            "Protectors At The Plate Autographed Relics": [("Red", 1)],
            "Dual Autograph Relic Collection": [("Black", 10), ("Red", 5), ("Gold", 1)],
        },
    },
    "2025-26-topps-transcendent": {
        "source_url": "https://baseballcardpedia.com/index.php/2025-26_Topps_Transcendent#Checklist",
        "sections": {
            "Base": [("Orange Refractor", 25), ("Gold Refractor", 10), ("Red Refractor", 5), ("SuperFractor", 1)],
            "Base Set - Image Variations": [("Orange Refractor", 25), ("Gold Refractor", 10), ("Red Refractor", 5), ("SuperFractor", 1)],
            "Transcendent Collection Legendary Relic Cards": [("Platinum", 1)],
            "Transcendent Icons Chrome Autograph": [("Blue", 10), ("Emerald", 5), ("Platinum", 1)],
            "Transcendent Collection Rookie Showcase Autographs": [("Blue", 10), ("Emerald", 5), ("Platinum", 1)],
            "MLB Logo Autograph Patch Card": [("Black Refractor", 10), ("Red Refractor", 5), ("SuperFractor", 1)],
            "Transcendent Autograph Relic": [("Blue", 10), ("Emerald", 5), ("Platinum", 1)],
            "Transcendent Collection Patch Autographs": [("Blue", 10), ("Emerald", 5), ("Platinum", 1)],
            "Transcendent Collection Dual Patch Autograph Variation": [("Emerald", 5), ("Platinum", 1)],
            "Transcendent Collection Rookie Showcase Patch Autographs": [("Platinum", 1)],
            "Transcendent Collection Jumbo Patch Autographs": [("Platinum", 1)],
            "Record Breaking Autograph Relic": [("MLB Logo Patch", 1)],
            "Transcendent Collection World Series Relic Autograph": [("Blue", 10), ("Emerald", 5), ("Platinum", 1)],
            "Transcendent Collection World Series Patch Autograph": [("Blue", 10), ("Emerald", 5), ("Platinum", 1)],
            "Transcendent Collection World Series Logo Patch Autograph": [("Platinum", 1)],
        },
    },
    "2026-bowman": {
        "source_url": "https://baseballcardpedia.com/index.php/2026_Bowman#Insertion_Ratios",
        "wiki_table": "https://baseballcardpedia.com/index.php?title=2026_Bowman&action=raw",
        "default_base_subset": "Base",
        "default_base_total": 100,
        "default_base_skip_prefixes": ["Mega Chrome", "Rookie Red RC"],
        "table_aliases": {
            "Base Set": "Base",
            "Prospects": "Base Prospects",
            "Chrome Prospects": "Chrome Prospects",
            "Etched in Glass": "Base - Etched In Glass Variations",
            "Chrome Prospects Etched in Glass": "Chrome Prospects - Etched In Glass Variations",
            "PackFractor": "Chrome Prospects Packfractor Variation",
            "Red RC Icon": "Base - Red RC Variations",
            "Scouts' Top 100": "Bowman Scouts Top 100",
            "Sterling": "Bowman Sterling",
            "Electric Sluggers": "Electric Sluggers",
            "Under the Radar": "Under The Radar",
            "Power Chords": "Power Chords",
            "Chrome Rookie Autographs": "Chrome Rookie Autographs",
            "Chrome Prospect Autographs": "Chrome Prospect Autographs",
            "Chrome Prospect Gold Ink Autographs": "Chrome Prospect Gold Ink Autographs",
            "Chrome Prospect PackFractor Autographs": "Chrome Prospect Packfractor Autographs",
            "Sterling Autographs": "Bowman Sterling Autographs",
            "Electric Sluggers Autographs": "Electric Sluggers Autographs",
            "Under the Radar Autographs": "Under The Radar Autographs",
            "Power Chords Autographs": "Power Chords Autographs",
        },
    },
    "2026-topps-series-2": {
        "source_url": "https://baseballcardpedia.com/index.php/2026_Topps#Insertion_Ratios",
        "wiki_table": "https://baseballcardpedia.com/index.php?title=2026_Topps&action=raw",
        "start_heading": "==Series Two==",
        "end_heading": "=Checklist=",
        "default_base_subset": "Base",
        "default_base_total": 350,
        "default_base_skip_prefixes": [
            "Team Color Border",
            "True Photo",
            "Golden Mirror",
            "Holiday",
            "Vintage Stock",
            "Clear",
            "Apple Foil",
            "1991",
        ],
        "table_aliases": {
            "Base": "Base",
            "Team Color Border Variation": "Team Color Border Variations",
            "True Photo Variation": "True Photo Variations",
            "Golden Mirror Variation": "Golden Mirror Variations",
            "Holiday Variation": "Holiday Variations",
            "Vintage Stock Variation": "Vintage Stock Variations",
            "Clear Variation": "Clear Variation",
            "Apple Foil Variation": "Apple Foil Variations",
            "1991 Topps Baseball": "1991 Topps Baseball",
            "1991 Topps Baseball Chrome": "1991 Topps Baseball Chrome",
            "1991 Topps All-Stars": "1991 Topps All-Stars",
            "1991 Topps Baseball Chrome All-Stars": "1991 Topps Baseball Chrome All-Stars",
            "The Flagship Collection": "The Flagship Collection",
            "The Flagship Collection Chrome": "The Flagship Collection - Chrome",
            "Stars of MLB": "Stars Of MLB",
            "All Aces": "All Aces",
            "All Kings": "All Kings",
            "Crooked Numbers": "Crooked Numbers",
            "Glove Work": "Glove Work",
            "Heavy Lumber": "Heavy Lumber",
            "Home Field Advantage": "Home Field",
            "Cover Athletes": "Cover Athletes",
            "Titans of the Game": "Titans Of The Game",
            "Swinging with the Stars": "Swinging With The Stars",
            "Diamond Dust": "Diamond Dust",
            "Topps Flagship Autograph Patches": "Topps Flagship Autograph Patches",
            "Baseball Stars Autographs": "Baseball Stars Autographs",
            "First Pitch Autographs": "First Pitch Autographs",
            "1991 Topps Baseball Autographs": "1991 Topps Baseball Autographs",
            "1991 Topps Baseball Chrome Autographs": "1991 Topps Baseball Chrome Autographs",
            "1991 Topps Baseball All-Star Autographs": "1991 Topps Baseball All-Star Autographs",
            "1991 Topps Baseball Chrome All-Star Autographs": "1991 Topps Baseball Chrome All-Star Autographs",
            "Major League Materials Autographs": "Major League Materials Autographs",
            "Major League Material": "Major League Material",
            "City Connect Swatch Collection": "City Connect Swatch Collection",
            "City Connect Swatches Autographs": "City Connect Swatches Autographs",
            "Real One Relics": "Real One Relics",
            "Rounding the Bases Relics": "Rounding The Bases Relics",
            "Rounding the Bases Autographs": "Rounding The Bases Autographs",
            "In the Name Relics": "In The Name Relics",
        },
    },
    "2026-topps-chrome-black": {
        "source_url": "https://baseballcardpedia.com/index.php/2026_Topps_Chrome_Black#Insertion_Ratios",
        "sections": {
            "Base": [
                ("Refractor", 199),
                ("Blue Refractor", 150),
                ("Green Refractor", 99),
                ("Purple Refractor", 75),
                ("Purple Mini-Diamond Refractor", 75),
                ("Purple Wave Refractor", 75),
                ("Gold Refractor", 50),
                ("Gold Mini Diamond Refractor", 50),
                ("Gold Wave Refractor", 50),
                ("Orange Refractor", 25),
                ("Orange Mini Diamond Refractor", 25),
                ("Orange Wave Refractor", 25),
                ("Rose Gold Refractor", 10),
                ("Rose Gold Mini Diamond Refractor", 10),
                ("Rose Gold Wave Refractor", 10),
                ("Red Refractor", 5),
                ("Red Mini Diamond Refractor", 5),
                ("Red Wave Refractor", 5),
                ("SuperFractor", 1),
            ],
            "Base - Rookie Design Variations": [
                ("Purple Refractor", 75),
                ("Gold Refractor", 50),
                ("Orange Refractor", 25),
                ("Rose Gold Refractor", 10),
                ("Red Refractor", 5),
                ("SuperFractor", 1),
            ],
            "Damascus": [("SuperFractor", 1)],
            "Nocturnal": [("SuperFractor", 1)],
            "Home Field": [("Black Refractor", 1)],
            "Depth Of Darkness": [
                ("Red Refractor", 5),
                ("SuperFractor", 1),
            ],
            "Chrome Black Autographs": [
                ("Refractor", 199),
                ("Blue Refractor", 150),
                ("Green Refractor", 99),
                ("Purple Refractor", 75),
                ("Gold Refractor", 50),
                ("Gold Mini Diamond Refractor", 50),
                ("Orange Refractor", 25),
                ("Orange Mini Diamond Refractor", 25),
                ("Rose Gold Refractor", 10),
                ("Rose Gold Mini Diamond Refractor", 10),
                ("Red Refractor", 5),
                ("Red Mini Diamond Refractor", 5),
                ("SuperFractor", 1),
            ],
            "Super Futures Autographs": [
                ("Gold Refractor", 50),
                ("Orange Refractor", 25),
                ("Rose Gold Refractor", 10),
                ("Red Refractor", 5),
                ("SuperFractor", 1),
            ],
            "Ivory Autographs": [
                ("Orange Trim", 25),
                ("Red Trim", 5),
                ("SuperFractor", 1),
            ],
            "Paint It": [
                ("Red Refractor", 5),
                ("SuperFractor", 1),
            ],
            "Pitch Black Pairings Dual Autographs": [
                ("Orange Refractor", 25),
                ("Rose Gold Refractor", 10),
                ("Red Refractor", 5),
                ("SuperFractor", 1),
            ],
        },
    },
    "2026-donruss": {
        "source_url": "https://baseballcardpedia.com/index.php/2026_Donruss#Checklist",
        "master_xlsx": "import/2026-Donruss-Baseball-Checklist.xlsx",
        "master_team": True,
    },
    "2026-panini-prizm-stars-stripes": {
        "source_url": "https://baseballcardpedia.com/index.php/2026_Panini_Stars_%26_Stripes_Prizm#Checklist",
        "master_xlsx": "import/2026-Panini-Prizm-Stars-Stripes-Baseball-Checklist.xlsx",
        "master_default_team": "USA Baseball",
    },
}


FIELD_ORDER = [
    "number",
    "uuid",
    "genre",
    "sport",
    "sports",
    "set_id",
    "subjects",
    "subset",
    "card_name",
    "description",
    "series",
    "variation",
    "parallel",
    "print_run",
    "rookie_card",
    "autograph",
    "relic",
    "serial_numbered",
    "release_date",
    "image_url",
    "external_links",
]


def slugify(value):
    value = str(value).lower().replace("&", "and")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def normalize(value):
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def print_run_from_text(value):
    value = str(value).strip().lower().replace(",", "")
    if not value or value in {"-", "n/a", "?"}:
        return None
    if "one-of-one" in value or value == "1/1":
        return 1
    if "four for each" in value:
        return 4
    words = {"ten": 10, "fifteen": 15, "five": 5}
    if value in words:
        return words[value]
    match = re.search(r"\d+", value)
    return int(match.group(0)) if match else None


def yaml_quote(value):
    escaped = str(value).replace("\\", "\\\\").replace('"', '\\"')
    return f'"{escaped}"'


def read_yaml(path):
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def read_existing_uuid(path):
    if not path.exists():
        return None
    match = re.search(r'^uuid:\s*"?([0-9a-fA-F-]{36})"?\s*$', path.read_text(encoding="utf-8"), re.MULTILINE)
    return match.group(1) if match else None


def write_card(path, data):
    lines = []
    for field in FIELD_ORDER:
        if field not in data:
            continue
        value = data[field]
        if value is None:
            continue
        if field == "subjects":
            lines.append("subjects:")
            for subject in value:
                lines.append(f"  - name: {yaml_quote(subject['name'])}")
                if "role" in subject:
                    lines.append(f"    role: {yaml_quote(subject['role'])}")
                if "team" in subject:
                    lines.append(f"    team: {yaml_quote(subject['team'])}")
        elif field == "external_links":
            lines.append("external_links:")
            for link in value:
                lines.append(f"  - name: {yaml_quote(link['name'])}")
                lines.append(f"    url: {yaml_quote(link['url'])}")
        elif isinstance(value, bool):
            lines.append(f"{field}: {str(value).lower()}")
        elif isinstance(value, int):
            lines.append(f"{field}: {value}")
        elif isinstance(value, list):
            lines.append(f"{field}: [{', '.join(yaml_quote(item) for item in value)}]")
        else:
            lines.append(f"{field}: {yaml_quote(value)}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def section_name(card):
    return card.get("subset", "Base")


def build_parallel_card(card, base_path, parallel, print_run, source_url):
    variant_path = base_path.with_name(f"{base_path.stem}-{slugify(parallel)}.yaml")
    variant = dict(card)
    variant["uuid"] = read_existing_uuid(variant_path) or str(uuid.uuid4())
    variant["parallel"] = parallel
    variant["print_run"] = print_run
    variant["serial_numbered"] = print_run is not None
    title = card.get("card_name", card["number"])
    if f" - {parallel}" not in title:
        variant["card_name"] = f"{title} - {parallel}"
    variant["description"] = f"{parallel} parallel of {card.get('description', title)}"
    image_root = card.get("image_url", "").rsplit("/", 1)[0]
    if image_root:
        variant["image_url"] = f"{image_root}/{variant_path.stem}.jpg"
    links = list(card.get("external_links", []))
    if not any(link.get("name") == "BaseballCardpedia" and link.get("url") == source_url for link in links):
        links.append({"name": "BaseballCardpedia", "url": source_url})
    variant["external_links"] = links
    return variant_path, variant


def update_set_count(set_dir):
    set_path = set_dir / "set.yaml"
    text = set_path.read_text(encoding="utf-8")
    count = len(list((set_dir / "cards").glob("*.yaml")))
    text = re.sub(r"^card_count:\s*\d+\s*$", f"card_count: {count}", text, flags=re.MULTILINE)
    set_path.write_text(text, encoding="utf-8")


def master_rows(plan):
    wb = load_workbook(plan["master_xlsx"], read_only=True, data_only=True)
    ws = wb["Master Card List"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        if plan.get("master_team"):
            card_set, number, athlete, team, sequence = (list(row) + [None] * 5)[:5]
        else:
            card_set, number, athlete, sequence = (list(row) + [None] * 4)[:4]
            team = plan.get("master_default_team", "")
        yield {
            "card_set": str(card_set).strip(),
            "number": str(number).strip(),
            "athlete": str(athlete).strip().rstrip(","),
            "team": str(team).strip() if team is not None else "",
            "print_run": int(sequence) if isinstance(sequence, int) else None,
        }


def existing_base_cards(cards_dir):
    cards = []
    for path in sorted(cards_dir.glob("*.yaml")):
        card = read_yaml(path)
        if "parallel" in card:
            continue
        cards.append((path, card, section_name(card), normalize(section_name(card))))
    return cards


def best_base_match(row, base_cards):
    row_set = normalize(row["card_set"])
    row_number = str(row["number"])
    candidates = []
    for path, card, subset, normalized_subset in base_cards:
        if str(card.get("number")) != row_number:
            continue
        if row_set == normalized_subset or row_set.startswith(normalized_subset):
            candidates.append((len(normalized_subset), path, card, subset))
    if not candidates:
        return None
    _, path, card, subset = max(candidates, key=lambda item: item[0])
    return path, card, subset


def master_parallel_name(card_set, subset):
    normalized_subset = normalize(subset)
    words = re.findall(r"[A-Za-z0-9]+", card_set)
    consumed = []
    for index in range(1, len(words) + 1):
        if normalize(" ".join(words[:index])) == normalized_subset:
            consumed = words[:index]
    if not consumed:
        return ""
    return " ".join(words[len(consumed) :]).strip()


def expand_master_set(set_id, limit):
    plan = PARALLEL_PLANS[set_id]
    set_dir = DATA_DIR / set_id
    cards_dir = set_dir / "cards"
    base_cards = existing_base_cards(cards_dir)
    written = 0
    for row in master_rows(plan):
        match = best_base_match(row, base_cards)
        if not match:
            continue
        base_path, card, subset = match
        parallel = master_parallel_name(row["card_set"], subset)
        if not parallel:
            continue
        variant_path, variant = build_parallel_card(card, base_path, parallel, row["print_run"], plan["source_url"])
        if variant_path.exists():
            continue
        write_card(variant_path, variant)
        written += 1
        if written >= limit:
            update_set_count(set_dir)
            return written
    update_set_count(set_dir)
    return written


def wiki_table_rows(plan):
    text = requests.get(plan["wiki_table"], timeout=30).text
    if plan.get("start_heading"):
        start = text.find(plan["start_heading"])
        if start != -1:
            text = text[start:]
    if plan.get("end_heading"):
        end = text.find(plan["end_heading"])
        if end != -1:
            text = text[:end]
    for line in text.splitlines():
        if not line.startswith("|"):
            continue
        cells = [cell.strip() for cell in line.strip("|").split("||")]
        if len(cells) < 3:
            continue
        label = re.sub(r"''+", "", cells[0]).strip()
        label = re.sub(r"\[\[(?:[^|\]]+\|)?([^\]]+)\]\]", r"\1", label)
        total = print_run_from_text(cells[1])
        print_run = print_run_from_text(cells[2])
        if not label or label.lower() == "cards":
            continue
        yield label, total, print_run


def wiki_variants(plan):
    aliases = sorted(plan["table_aliases"].items(), key=lambda item: len(normalize(item[0])), reverse=True)
    variants = {}
    for label, total, print_run in wiki_table_rows(plan):
        normalized_label = normalize(label)
        matched = False
        for prefix, subset in aliases:
            normalized_prefix = normalize(prefix)
            if normalized_label == normalized_prefix:
                matched = True
                break
            if normalized_label.startswith(normalized_prefix):
                parallel = label[len(prefix) :].strip(" -")
                if not parallel or normalize(parallel) in {"s", "base", "set", "autograph", "autographs", "relic", "relics"}:
                    matched = True
                    break
                variants.setdefault(subset, [])
                item = (parallel, print_run)
                if item not in variants[subset]:
                    variants[subset].append(item)
                matched = True
                break
        if matched:
            continue
        if plan.get("default_base_subset") and total == plan.get("default_base_total"):
            skip = any(normalized_label.startswith(normalize(prefix)) for prefix in plan.get("default_base_skip_prefixes", []))
            if not skip and normalized_label not in {"base", "baseset"}:
                variants.setdefault(plan["default_base_subset"], [])
                item = (label, print_run)
                if item not in variants[plan["default_base_subset"]]:
                    variants[plan["default_base_subset"]].append(item)
    return variants


def expand_set(set_id, limit):
    plan = PARALLEL_PLANS[set_id]
    if "master_xlsx" in plan:
        return expand_master_set(set_id, limit)
    if "wiki_table" in plan:
        plan = dict(plan)
        plan["sections"] = wiki_variants(plan)
    set_dir = DATA_DIR / set_id
    cards_dir = set_dir / "cards"
    written = 0
    for path in sorted(cards_dir.glob("*.yaml")):
        card = read_yaml(path)
        if "parallel" in card:
            continue
        variants = plan["sections"].get(section_name(card), [])
        for parallel, print_run in variants:
            variant_path, variant = build_parallel_card(card, path, parallel, print_run, plan["source_url"])
            if variant_path.exists():
                continue
            write_card(variant_path, variant)
            written += 1
            if written >= limit:
                update_set_count(set_dir)
                return written
    update_set_count(set_dir)
    return written


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--set-id", choices=sorted(PARALLEL_PLANS), required=True)
    parser.add_argument("--limit", type=int, default=1000)
    return parser.parse_args()


def main():
    args = parse_args()
    written = expand_set(args.set_id, args.limit)
    print(f"Wrote {written} parallel cards for {args.set_id}")


if __name__ == "__main__":
    main()

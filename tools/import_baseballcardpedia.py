import argparse
import html
import re
import uuid
from html.parser import HTMLParser
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook


DATA_DIR = Path("data")
SERIES = "Topps Chrome Platinum Anniversary"
SET_NAME = "2025-26 Topps Chrome Platinum Anniversary"
RELEASE_DATE = "2026-06-05"
PLACEHOLDER_IMAGE_ROOT = "https://example.com/cards/2025-26-topps-chrome-platinum"


class DescriptionParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.in_heading = False
        self.in_description = False
        self.current_heading = None
        self.heading_parts = []
        self.description_parts = []
        self.descriptions = []

    def handle_starttag(self, tag, attrs):
        if tag in {"h1", "h2", "h3"}:
            self.in_heading = True
            self.heading_parts = []
        elif tag == "p" and self.current_heading == "Description":
            self.in_description = True
            self.description_parts = []

    def handle_endtag(self, tag):
        if self.in_heading and tag in {"h1", "h2", "h3"}:
            self.current_heading = clean_text("".join(self.heading_parts))
            self.in_heading = False
        elif tag == "p" and self.in_description:
            text = clean_text("".join(self.description_parts))
            if text:
                self.descriptions.append(text)
            self.in_description = False

    def handle_data(self, data):
        if self.in_heading:
            self.heading_parts.append(data)
        if self.in_description:
            self.description_parts.append(data)

    def handle_entityref(self, name):
        self.handle_data(html.unescape(f"&{name};"))

    def handle_charref(self, name):
        self.handle_data(html.unescape(f"&#{name};"))


SECTION_NAMES = {
    "Base": "Base",
    "Base - Image Variations": "Base - Image Variations",
    "Base - City Variations": "Base - City Variations",
    "Chrome Platinum Autographs": "Chrome Platinum Autographs",
    "1955 Topps City Variations Autographs": "1955 Topps City Variations Autographs",
    "1955 World Series": "1955 World Series",
    "1955 Topps Rails And Sails": "1955 Topps Rails And Sails",
    "1955 Topps Doubleheaders": "1955 Topps Doubleheaders",
    "1955 Cards That Never Were": "1955 Cards That Never Were",
    "1955 Topps Employee Super Short Prints": "1955 Topps Employee Super Short Prints",
}


SHEETS = ["Base", "Variations", "Autographs", "Inserts"]


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", html.unescape(str(value))).strip()


def clean_name(value):
    return clean_text(value).rstrip(",").strip()


def slugify(value):
    value = str(value).lower().replace("&", "and")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def yaml_quote(value):
    escaped = str(value).replace("\\", "\\\\").replace('"', '\\"')
    return f'"{escaped}"'


def read_existing_uuid(path):
    if not path.exists():
        return None
    match = re.search(r'^uuid:\s*"?([0-9a-fA-F-]{36})"?\s*$', path.read_text(encoding="utf-8"), re.MULTILINE)
    return match.group(1) if match else None


def get_uuid(path):
    return read_existing_uuid(path) or str(uuid.uuid4())


def card_filename(number, section):
    prefix = "" if section == "Base" else f"{slugify(section)}-"
    safe = re.sub(r"[^A-Za-z0-9_-]+", "-", f"{prefix}{number}").strip("-")
    return f"{safe}.yaml"


def split_subjects(name):
    return [part.strip() for part in re.split(r"\s*/\s*", clean_name(name)) if part.strip()]


def is_rookie(value):
    return clean_text(value).upper() == "RC"


def is_card_row(values):
    number, name, team = values[:3]
    return number not in (None, "") and name not in (None, "") and team not in (None, "")


def is_section_header(values):
    first = clean_text(values[0] if values else None)
    return first in SECTION_NAMES


def read_cards_from_workbook(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    cards = []
    for sheet_name in SHEETS:
        ws = wb[sheet_name]
        current_section = None
        for row in ws.iter_rows(values_only=True):
            values = list(row) + [None] * 5
            if is_section_header(values):
                current_section = SECTION_NAMES[clean_text(values[0])]
                continue
            if current_section and is_card_row(values):
                cards.append(
                    {
                        "number": str(values[0]).strip(),
                        "names": split_subjects(values[1]),
                        "team": clean_text(values[2]),
                        "rookie_card": is_rookie(values[3]),
                        "section": current_section,
                    }
                )
    return cards


def fetch_description(url):
    request = Request(url, headers={"User-Agent": "open-checklist-project-importer/0.2"})
    with urlopen(request, timeout=30) as response:
        parser = DescriptionParser()
        parser.feed(response.read().decode("utf-8", errors="replace"))
    return parser.descriptions[0] if parser.descriptions else f"{SET_NAME} checklist."


def card_description(card):
    name = " / ".join(card["names"])
    section = card["section"]
    if section == "Base":
        return f"Base card featuring {name}."
    if "Autograph" in section:
        return f"Autographed {section} card featuring {name}."
    return f"{section} card featuring {name}."


def card_name(card):
    name = " / ".join(card["names"])
    return f"{name} - Base" if card["section"] == "Base" else f"{name} - {card['section']}"


def variation_value(section):
    if section == "Base - Image Variations":
        return "Image Variation"
    if section == "Base - City Variations":
        return "City Variation"
    if section == "1955 Topps City Variations Autographs":
        return "City Variation Autograph"
    return None


def write_set(set_dir, set_id, source_url, description, cards):
    set_path = set_dir / "set.yaml"
    counts = {}
    for card in cards:
        counts[card["section"]] = counts.get(card["section"], 0) + 1

    lines = [
        f"uuid: {get_uuid(set_path)}",
        f"set_id: {yaml_quote(set_id)}",
        f"name: {yaml_quote(SET_NAME)}",
        'genre: "Sports"',
        'category: ["MLB"]',
        'sports: ["Baseball"]',
        'season: "2025-26"',
        "years: [2025, 2026]",
        f"series: {yaml_quote(SERIES)}",
        f"release_date: {yaml_quote(RELEASE_DATE)}",
        'manufacturer: "Topps"',
        f"card_count: {len(cards)}",
        f"description: {yaml_quote(description)}",
        f"image_url: {yaml_quote('https://example.com/images/2025-26-topps-chrome-platinum.jpg')}",
        "metadata:",
        '  language: "en"',
        "  year: 2026",
        f"  source_name: {yaml_quote('BaseballCardpedia')}",
        f"  source_url: {yaml_quote(source_url)}",
        "  source_file: \"import/2025-Topps-Chrome-Platinum-Baseball-Checklist.xlsx\"",
        "  subsets:",
    ]
    for section, count in counts.items():
        lines.append(f"    - name: {yaml_quote(section)}")
        lines.append(f"      card_count: {count}")

    set_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_cards(set_dir, set_id, cards):
    cards_dir = set_dir / "cards"
    cards_dir.mkdir(parents=True, exist_ok=True)

    for card in cards:
        path = cards_dir / card_filename(card["number"], card["section"])
        autograph = "Autograph" in card["section"]
        variation = variation_value(card["section"])
        role = "Employee" if "Employee" in card["section"] else "Player"
        image_slug = path.stem

        lines = [
            f"number: {yaml_quote(card['number'])}",
            f"uuid: {yaml_quote(get_uuid(path))}",
            'genre: "Sports"',
            'sport: "Baseball"',
            f"set_id: {yaml_quote(set_id)}",
            "subjects:",
        ]
        for name in card["names"]:
            lines.append(f"  - name: {yaml_quote(name)}")
            lines.append(f"    role: {yaml_quote(role)}")
            lines.append(f"    team: {yaml_quote(card['team'])}")

        if card["section"] != "Base":
            lines.append(f"subset: {yaml_quote(card['section'])}")
        lines.extend(
            [
                f"card_name: {yaml_quote(card_name(card))}",
                f"description: {yaml_quote(card_description(card))}",
                f"series: {yaml_quote(SERIES)}",
            ]
        )
        if variation:
            lines.append(f"variation: {yaml_quote(variation)}")
        lines.extend(
            [
                f"rookie_card: {str(card['rookie_card']).lower()}",
                f"autograph: {str(autograph).lower()}",
                "serial_numbered: false",
                f"release_date: {yaml_quote(RELEASE_DATE)}",
                f"image_url: {yaml_quote(f'{PLACEHOLDER_IMAGE_ROOT}/{image_slug}.jpg')}",
            ]
        )

        path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("url")
    parser.add_argument("--xlsx", default="import/2025-Topps-Chrome-Platinum-Baseball-Checklist.xlsx")
    parser.add_argument("--set-id", default="2025-26-topps-chrome-platinum")
    parser.add_argument("--category", default="baseball")
    args = parser.parse_args()

    cards = read_cards_from_workbook(Path(args.xlsx))
    set_dir = DATA_DIR / args.category / args.set_id
    set_dir.mkdir(parents=True, exist_ok=True)
    write_set(set_dir, args.set_id, args.url, fetch_description(args.url), cards)
    write_cards(set_dir, args.set_id, cards)
    print(f"Wrote {len(cards)} cards to {set_dir}")


if __name__ == "__main__":
    main()
